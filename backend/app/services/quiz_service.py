"""安全测验领域服务（UC-05 / UC-09）。

职责
----
本模块覆盖三个使用面：

- **管理员题库管理**：增删改查题目（``QUIZ_BANK_MANAGE``）。
- **管理员指定测验**：发起 ASSIGNED 测验、撤回、完成率报告、Excel 导出。
- **学生答题**：随机练习（``RANDOM``）/ 指定测验（``ASSIGNED``）的开测与提交，
  错题自动关联知识库条目用于推送学习。

约定
----
- 所有事务边界用 ``async with uow():``，仓储不主动 commit。
- 控制层不抛 ``HTTPException``，业务异常一律抛 :mod:`app.exceptions` 子类。
- ID 在 schema 出口处转 ``str``，避免 JS 安全整数边界。
"""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging import get_logger
from app.core.snowflake import next_snowflake_id
from app.domain.user_snapshot import UserSnapshot
from app.exceptions import (
    PermissionDenied,
    QuestionNotFound,
    QuizAlreadySubmitted,
    QuizClosed,
    QuizInvalidParam,
    QuizNotEligible,
    QuizNotFound,
    QuizQuestionBankExhausted,
)
from app.infra.db.models import Department, QuestionBank, Quiz, User
from app.infra.db.models.question_bank import QuestionDifficulty
from app.infra.db.models.quiz import QuizScopeType, QuizStatus, QuizType
from app.infra.db.models.quiz_attempt import QuizAttempt, QuizAttemptStatus
from app.infra.db.models.quiz_attempt_answer import QuizAttemptAnswer
from app.infra.db.models.role import Role
from app.infra.db.models.user import UserStatus
from app.infra.db.session import uow
from app.infra.repositories.quiz import QuizRepository
from app.schemas.common import PaginationOut
from app.schemas.quiz import (
    AssignedQuizCreateIn,
    DepartmentCompletionItem,
    QuestionAdminOut,
    QuestionCreateIn,
    QuestionResultOut,
    QuestionStudentOut,
    QuestionUpdateIn,
    QuizCancelIn,
    QuizCompletionReportOut,
    QuizDetailOut,
    QuizHistoryItemOut,
    QuizListItemOut,
    StartQuizOut,
    SubmitQuizIn,
    SubmitQuizOut,
    WrongQuestionOut,
)
from app.services.audit_service import get_audit_service
from app.services.notification_service import send_notification

logger = get_logger(__name__)

# 随机练习默认抽题数（与 schemas.quiz 中 ``Quiz.question_count`` 默认一致）
RANDOM_QUIZ_QUESTION_COUNT = 10
RANDOM_QUIZ_PASS_SCORE = 60


# ╔══════════════════════════════════════════════════════════════════╗
# ║ 1. 管理员：题库管理 (UC-05)                                       ║
# ╚══════════════════════════════════════════════════════════════════╝


async def create_question(
    body: QuestionCreateIn, *, current: UserSnapshot
) -> QuestionAdminOut:
    """新增题目。"""
    async with uow() as session:
        repo = QuizRepository(session)
        q = QuestionBank(
            question_id=next_snowflake_id(),
            content=body.content,
            option_a=body.option_a,
            option_b=body.option_b,
            option_c=body.option_c,
            option_d=body.option_d,
            correct_answer=body.correct_answer,
            explanation=body.explanation,
            fraud_type_id=body.fraud_type_id,
            knowledge_entry_id=body.knowledge_entry_id,
            difficulty=body.difficulty,
            is_active=1,
            created_by=current.user_id,
        )
        await repo.add_question(q)
        await get_audit_service().write(
            operator=current,
            op_type="QUIZ_QUESTION_CREATE",
            obj_type="question_bank",
            obj_id=str(q.question_id),
            after={"content": body.content[:80], "difficulty": body.difficulty},
            sync=True,
            session=session,
        )
        out = _question_admin_out(q)
    logger.info("quiz_question_created", question_id=q.question_id, by=current.user_id)
    return out


async def update_question(
    question_id: int, body: QuestionUpdateIn, *, current: UserSnapshot
) -> QuestionAdminOut:
    """编辑题目（PATCH）。"""
    async with uow() as session:
        repo = QuizRepository(session)
        q = await repo.get_question(question_id)
        if q is None:
            raise QuestionNotFound()
        before = {
            "content": q.content,
            "correct_answer": q.correct_answer,
            "is_active": bool(q.is_active),
            "difficulty": q.difficulty,
        }
        data = body.model_dump(exclude_unset=True)
        is_active = data.pop("is_active", None)
        for k, v in data.items():
            setattr(q, k, v)
        if is_active is not None:
            q.is_active = 1 if is_active else 0
        await session.flush()
        await get_audit_service().write(
            operator=current,
            op_type="QUIZ_QUESTION_UPDATE",
            obj_type="question_bank",
            obj_id=str(q.question_id),
            before=before,
            after={"changes": list(data.keys()), "is_active": q.is_active},
            sync=True,
            session=session,
        )
        out = _question_admin_out(q)
    return out


async def delete_question(question_id: int, *, current: UserSnapshot) -> dict[str, str]:
    """软删除（``is_active=0``）；保留历史答题记录的外键完整性。"""
    async with uow() as session:
        repo = QuizRepository(session)
        q = await repo.get_question(question_id)
        if q is None:
            raise QuestionNotFound()
        if q.is_active == 0:
            return {"question_id": str(question_id), "status": "already_inactive"}
        q.is_active = 0
        await session.flush()
        await get_audit_service().write(
            operator=current,
            op_type="QUIZ_QUESTION_DELETE",
            obj_type="question_bank",
            obj_id=str(question_id),
            sync=True,
            session=session,
        )
    return {"question_id": str(question_id), "status": "inactive"}


async def list_questions(
    *,
    keyword: str | None = None,
    fraud_type_id: int | None = None,
    difficulty: int | None = None,
    is_active: bool | None = None,
    page: int = 1,
    size: int = 20,
) -> PaginationOut[QuestionAdminOut]:
    """管理员题库分页。"""
    if difficulty is not None and difficulty not in QuestionDifficulty.ALL:
        raise QuizInvalidParam("难度参数非法")
    offset = (page - 1) * size
    async with uow() as session:
        repo = QuizRepository(session)
        rows, total = await repo.list_questions(
            keyword=keyword,
            fraud_type_id=fraud_type_id,
            difficulty=difficulty,
            is_active=is_active,
            offset=offset,
            limit=size,
        )
        items = [_question_admin_out(q) for q in rows]
    return PaginationOut[QuestionAdminOut](items=items, total=total, page=page, size=size)


async def get_question(question_id: int) -> QuestionAdminOut:
    async with uow() as session:
        repo = QuizRepository(session)
        q = await repo.get_question(question_id)
        if q is None:
            raise QuestionNotFound()
        return _question_admin_out(q)


# ╔══════════════════════════════════════════════════════════════════╗
# ║ 2. 管理员：发起 / 撤回指定测验 (UC-09)                            ║
# ╚══════════════════════════════════════════════════════════════════╝


async def create_assigned_quiz(
    body: AssignedQuizCreateIn, *, current: UserSnapshot
) -> QuizDetailOut:
    """管理员发起一场指定测验。"""
    now = datetime.now(tz=UTC)
    if body.deadline_at <= now:
        raise QuizInvalidParam("截止时间必须在当前时间之后")

    async with uow() as session:
        repo = QuizRepository(session)

        # 题目存在性 + 启用校验
        questions = await repo.get_questions_by_ids(list(set(body.question_ids)))
        if len(questions) != len(set(body.question_ids)):
            raise QuestionNotFound("存在未找到的题目 ID")
        if any(q.is_active == 0 for q in questions):
            raise QuizInvalidParam("题目列表中包含已禁用题目")

        # 参与范围校验
        is_dept_reviewer = await _is_dept_reviewer(session, current=current)
        if is_dept_reviewer:
            target_scope = {"type": "DEPT", "dept_ids": [current.department_id]}
        else:
            target_scope = await _build_target_scope(body, session=session)

        publish_level = 1 if is_dept_reviewer else 2

        quiz_id = next_snowflake_id()
        quiz = Quiz(
            quiz_id=quiz_id,
            quiz_type=QuizType.ASSIGNED,
            title=body.title,
            question_count=len(body.question_ids),
            pass_score=body.pass_score,
            status=QuizStatus.ACTIVE,
            created_by=current.user_id,
            deadline_at=body.deadline_at,
            target_scope=target_scope,
            publish_level=publish_level,
        )
        await repo.add_quiz(quiz)
        await repo.add_quiz_questions(
            quiz_id=quiz_id, ordered_question_ids=list(body.question_ids)
        )

        # 通知受众（学生 + ACTIVE）
        recipients = await _resolve_quiz_recipients(
            session, target_scope=quiz.target_scope or {}
        )
        for stu in recipients:
            await send_notification(
                recipient_id=stu.user_id,
                type="QUIZ_ASSIGNED",
                title=f"[安全测验] {body.title}",
                content=f"截止时间：{body.deadline_at.isoformat()}",
                related_object_type="quiz",
                related_object_id=quiz_id,
                db_session=session,
            )

        await get_audit_service().write(
            operator=current,
            op_type="QUIZ_ASSIGN",
            obj_type="quiz",
            obj_id=str(quiz_id),
            after={
                "title": body.title,
                "scope": target_scope.get("type"),
                "question_count": len(body.question_ids),
                "deadline_at": body.deadline_at.isoformat(),
                "recipient_count": len(recipients),
            },
            sync=True,
            session=session,
        )
        out = _quiz_detail_out(quiz)
    logger.info(
        "quiz_assigned",
        quiz_id=quiz_id,
        scope=target_scope.get("type"),
        recipient_count=len(recipients),
        operator=current.user_id,
    )
    return out


async def cancel_quiz(
    quiz_id: int, body: QuizCancelIn, *, current: UserSnapshot
) -> dict[str, str]:
    """撤回测验：将 ACTIVE 的指定测验改为 CANCELLED。"""
    async with uow() as session:
        repo = QuizRepository(session)
        quiz = await repo.get_quiz(quiz_id)
        if quiz is None or quiz.quiz_type != QuizType.ASSIGNED:
            raise QuizNotFound()
        await _ensure_admin_can_cancel_quiz(session, current=current, quiz=quiz)
        if quiz.status != QuizStatus.ACTIVE:
            raise QuizClosed("仅 ACTIVE 状态的测验可撤回")
        quiz.status = QuizStatus.CANCELLED
        await session.flush()
        await get_audit_service().write(
            operator=current,
            op_type="QUIZ_CANCEL",
            obj_type="quiz",
            obj_id=str(quiz_id),
            after={"reason": body.reason[:120]},
            sync=True,
            session=session,
        )
    logger.info("quiz_cancelled", quiz_id=quiz_id, operator=current.user_id)
    return {"quiz_id": str(quiz_id), "status": QuizStatus.CANCELLED}


async def list_assigned_quizzes_admin(
    *,
    current: UserSnapshot,
    status: str | None = None,
    keyword: str | None = None,
    page: int = 1,
    size: int = 20,
) -> PaginationOut[QuizListItemOut]:
    offset = (page - 1) * size
    async with uow() as session:
        repo = QuizRepository(session)
        rows, total = await repo.list_assigned_quizzes(
            status=status, keyword=keyword, offset=offset, limit=size
        )

        if await _is_school_reviewer(session, current=current):
            filtered = [q for q in rows if int(getattr(q, "publish_level", 2) or 2) != 1]
            items = [_quiz_list_item_out(q) for q in filtered]
            return PaginationOut[QuizListItemOut](
                items=items, total=len(filtered), page=page, size=size
            )

        if not await _is_dept_reviewer(session, current=current):
            items = [_quiz_list_item_out(q) for q in rows]
            return PaginationOut[QuizListItemOut](
                items=items, total=total, page=page, size=size
            )

        filtered: list[Quiz] = []
        for q in rows:
            try:
                await _ensure_admin_can_access_quiz(session, current=current, quiz=q)
            except PermissionDenied:
                continue
            filtered.append(q)

        items = [_quiz_list_item_out(q) for q in filtered]
        return PaginationOut[QuizListItemOut](
            items=items, total=len(filtered), page=page, size=size
        )


async def get_quiz_admin(quiz_id: int, *, current: UserSnapshot) -> QuizDetailOut:
    async with uow() as session:
        repo = QuizRepository(session)
        quiz = await repo.get_quiz(quiz_id)
        if quiz is None:
            raise QuizNotFound()
        scope = quiz.target_scope or {}
        await _ensure_admin_can_access_quiz(session, current=current, quiz=quiz)
        out = _quiz_detail_out(quiz)
        out.target_scope = scope
        return out


# ╔══════════════════════════════════════════════════════════════════╗
# ║ 3. 学生：随机练习 / 指定测验                                       ║
# ╚══════════════════════════════════════════════════════════════════╝


async def start_random_quiz(*, current: UserSnapshot) -> StartQuizOut:
    """学生主动发起一次随机练习（UC-05）。

    每次调用都新建一个 ``RANDOM`` 测验 + 题目快照，10 题随机。
    """
    async with uow() as session:
        repo = QuizRepository(session)
        active_count = await repo.count_active_questions()
        if active_count < RANDOM_QUIZ_QUESTION_COUNT:
            raise QuizQuestionBankExhausted(
                f"题库可用题目不足 {RANDOM_QUIZ_QUESTION_COUNT} 道，请联系管理员补题"
            )
        sampled = await repo.sample_active_questions(n=RANDOM_QUIZ_QUESTION_COUNT)

        quiz_id = next_snowflake_id()
        quiz = Quiz(
            quiz_id=quiz_id,
            quiz_type=QuizType.RANDOM,
            title="随机练习",
            question_count=len(sampled),
            pass_score=RANDOM_QUIZ_PASS_SCORE,
            status=QuizStatus.ACTIVE,
            created_by=current.user_id,
            publish_level=2,
            target_scope=None,
            deadline_at=None,
        )
        await repo.add_quiz(quiz)
        await repo.add_quiz_questions(
            quiz_id=quiz_id, ordered_question_ids=[q.question_id for q in sampled]
        )

        attempt = QuizAttempt(
            attempt_id=next_snowflake_id(),
            quiz_id=quiz_id,
            student_id=current.user_id,
            status=QuizAttemptStatus.IN_PROGRESS,
        )
        await repo.add_attempt(attempt)

        student_qs = [
            QuestionStudentOut(
                question_id=str(q.question_id),
                sort_order=idx,
                content=q.content,
                option_a=q.option_a,
                option_b=q.option_b,
                option_c=q.option_c,
                option_d=q.option_d,
            )
            for idx, q in enumerate(sampled, start=1)
        ]
    return StartQuizOut(
        quiz_id=str(quiz_id),
        attempt_id=str(attempt.attempt_id),
        title=quiz.title,
        pass_score=quiz.pass_score,
        question_count=quiz.question_count,
        questions=student_qs,
    )


async def start_assigned_quiz(
    quiz_id: int, *, current: UserSnapshot
) -> StartQuizOut:
    """学生开始一次指定测验（UC-09）。

    幂等性：同一 (quiz_id, student_id) 只允许有一条 attempt；若已存在且未提交，
    直接复用并返回题目；若已提交则抛 :class:`QuizAlreadySubmitted`。
    """
    async with uow() as session:
        repo = QuizRepository(session)
        quiz = await repo.get_quiz(quiz_id)
        if quiz is None or quiz.quiz_type != QuizType.ASSIGNED:
            raise QuizNotFound()
        await _ensure_quiz_open(quiz)
        await _ensure_student_in_scope(
            session, quiz=quiz, user_id=current.user_id
        )

        attempt = await repo.get_attempt_by_quiz_user(
            quiz_id=quiz_id, user_id=current.user_id
        )
        if attempt is not None and attempt.status == QuizAttemptStatus.SUBMITTED:
            raise QuizAlreadySubmitted()
        if attempt is None:
            attempt = QuizAttempt(
                attempt_id=next_snowflake_id(),
                quiz_id=quiz_id,
                student_id=current.user_id,
                status=QuizAttemptStatus.IN_PROGRESS,
            )
            await repo.add_attempt(attempt)

        links = await repo.list_quiz_questions(quiz_id)
        if not links:
            raise QuizInvalidParam("测验未关联任何题目，请联系管理员")
        questions = await repo.get_questions_by_ids([lk.question_id for lk in links])
        by_id = {q.question_id: q for q in questions}
        student_qs = [
            QuestionStudentOut(
                question_id=str(lk.question_id),
                sort_order=lk.sort_order,
                content=by_id[lk.question_id].content,
                option_a=by_id[lk.question_id].option_a,
                option_b=by_id[lk.question_id].option_b,
                option_c=by_id[lk.question_id].option_c,
                option_d=by_id[lk.question_id].option_d,
            )
            for lk in links
            if lk.question_id in by_id
        ]
    return StartQuizOut(
        quiz_id=str(quiz_id),
        attempt_id=str(attempt.attempt_id),
        title=quiz.title,
        pass_score=quiz.pass_score,
        question_count=quiz.question_count,
        questions=student_qs,
    )


async def submit_quiz(
    attempt_id: int, body: SubmitQuizIn, *, current: UserSnapshot
) -> SubmitQuizOut:
    """学生提交答卷。计分 + 写答题明细 + 返回每题正确答案与解析。"""
    async with uow() as session:
        repo = QuizRepository(session)
        attempt = await repo.get_attempt(attempt_id)
        if attempt is None or attempt.student_id != current.user_id:
            raise QuizNotFound("答题记录不存在")
        if attempt.status == QuizAttemptStatus.SUBMITTED:
            raise QuizAlreadySubmitted()

        quiz = await repo.get_quiz(attempt.quiz_id)
        if quiz is None:
            raise QuizNotFound()
        # ASSIGNED：再校验测验状态；RANDOM：不校验（学生自己发起的）
        if quiz.quiz_type == QuizType.ASSIGNED:
            await _ensure_quiz_open(quiz)

        # 取题目原始数据用于判分 + 推送知识库
        links = await repo.list_quiz_questions(quiz.quiz_id)
        ordered_qids = [lk.question_id for lk in links]
        if not ordered_qids:
            raise QuizInvalidParam("测验未关联任何题目")
        questions = await repo.get_questions_by_ids(ordered_qids)
        by_id: dict[int, QuestionBank] = {q.question_id: q for q in questions}

        # 标准化用户作答（按 question_id 索引）
        chosen_map: dict[int, str | None] = {
            item.question_id: item.chosen_answer for item in body.answers
        }

        # 计分 + 写明细
        correct = 0
        results: list[QuestionResultOut] = []
        answers_to_save: list[QuizAttemptAnswer] = []
        for qid in ordered_qids:
            q = by_id.get(qid)
            if q is None:
                # 题目被删（is_active=0 不影响外键存在）— 视为缺题，跳过
                continue
            chosen = chosen_map.get(qid)
            is_correct = chosen is not None and chosen == q.correct_answer
            if is_correct:
                correct += 1
            answers_to_save.append(
                QuizAttemptAnswer(
                    answer_id=next_snowflake_id(),
                    attempt_id=attempt.attempt_id,
                    question_id=qid,
                    chosen_answer=chosen,
                    is_correct=1 if is_correct else 0,
                )
            )
            results.append(
                QuestionResultOut(
                    question_id=str(q.question_id),
                    content=q.content,
                    option_a=q.option_a,
                    option_b=q.option_b,
                    option_c=q.option_c,
                    option_d=q.option_d,
                    correct_answer=q.correct_answer,
                    chosen_answer=chosen,
                    is_correct=is_correct,
                    explanation=q.explanation,
                    knowledge_entry_id=(
                        str(q.knowledge_entry_id)
                        if q.knowledge_entry_id is not None
                        else None
                    ),
                )
            )
        await repo.add_answers(answers_to_save)

        total = len(results)
        score = int(round(correct * 100 / total)) if total else 0
        attempt.score = score
        attempt.correct_count = correct
        attempt.status = QuizAttemptStatus.SUBMITTED
        attempt.submitted_at = datetime.now(tz=UTC)
        await session.flush()

        await get_audit_service().write(
            operator=current,
            op_type="QUIZ_SUBMIT",
            obj_type="quiz_attempt",
            obj_id=str(attempt.attempt_id),
            after={
                "quiz_id": str(quiz.quiz_id),
                "score": score,
                "correct": correct,
                "total": total,
                "quiz_type": quiz.quiz_type,
            },
            sync=True,
            session=session,
        )
    return SubmitQuizOut(
        quiz_id=str(quiz.quiz_id),
        attempt_id=str(attempt.attempt_id),
        score=score,
        pass_score=quiz.pass_score,
        is_pass=score >= quiz.pass_score,
        correct_count=correct,
        total_count=total,
        submitted_at=attempt.submitted_at,
        results=results,
    )


async def list_assigned_quizzes_for_student(
    *,
    current: UserSnapshot,
    status: str | None = None,
) -> list[QuizListItemOut]:
    """学生侧：本人可参与的所有指定测验（含本人作答状态）。"""
    async with uow() as session:
        repo = QuizRepository(session)
        quizzes = await repo.list_assigned_quizzes_for_student(
            dept_id=current.department_id,
            user_id=current.user_id,
            status=status,
        )
        out: list[QuizListItemOut] = []
        for q in quizzes:
            attempt = await repo.get_attempt_by_quiz_user(
                quiz_id=q.quiz_id, user_id=current.user_id
            )
            out.append(
                QuizListItemOut(
                    quiz_id=str(q.quiz_id),
                    quiz_type=q.quiz_type,
                    title=q.title,
                    question_count=q.question_count,
                    pass_score=q.pass_score,
                    status=q.status,
                    deadline_at=q.deadline_at,
                    created_at=q.created_at,
                    publish_level=int(getattr(q, "publish_level", 2) or 2),
                    my_attempt_status=attempt.status if attempt else None,
                    my_score=attempt.score if attempt else None,
                )
            )
    return out


async def list_wrong_questions_for_student(
    *, current: UserSnapshot, limit: int = 100
) -> list[WrongQuestionOut]:
    """错题汇总（用于错题本 / 跳转知识库学习）。"""
    async with uow() as session:
        repo = QuizRepository(session)
        rows = await repo.list_wrong_answers_for_student(
            user_id=current.user_id, limit=limit
        )
        if not rows:
            return []
        qids = list({a.question_id for a, _ in rows})
        questions = await repo.get_questions_by_ids(qids)
        by_id = {q.question_id: q for q in questions}
        out: list[WrongQuestionOut] = []
        for ans, attempt in rows:
            q = by_id.get(ans.question_id)
            if q is None:
                continue
            out.append(
                WrongQuestionOut(
                    question_id=str(q.question_id),
                    quiz_id=str(attempt.quiz_id),
                    attempt_id=str(attempt.attempt_id),
                    content=q.content,
                    option_a=q.option_a,
                    option_b=q.option_b,
                    option_c=q.option_c,
                    option_d=q.option_d,
                    correct_answer=q.correct_answer,
                    chosen_answer=ans.chosen_answer,
                    explanation=q.explanation,
                    knowledge_entry_id=(
                        str(q.knowledge_entry_id)
                        if q.knowledge_entry_id is not None
                        else None
                    ),
                    wrong_at=attempt.submitted_at or attempt.started_at,
                )
            )
    return out


async def list_quiz_history_for_student(
    *, current: UserSnapshot, page: int = 1, size: int = 20
) -> PaginationOut[QuizHistoryItemOut]:
    """测验历史记录（学生个人中心）。

    返回本人所有已提交答卷（含随机练习与指定测验），按提交时间倒序分页。
    """
    if page < 1:
        page = 1
    if size < 1 or size > 100:
        size = 20
    offset = (page - 1) * size
    async with uow() as session:
        repo = QuizRepository(session)
        rows, total = await repo.list_submitted_attempts_for_student(
            user_id=current.user_id, limit=size, offset=offset
        )
        items: list[QuizHistoryItemOut] = []
        for attempt, quiz in rows:
            score = int(attempt.score or 0)
            correct = int(attempt.correct_count or 0)
            items.append(
                QuizHistoryItemOut(
                    attempt_id=str(attempt.attempt_id),
                    quiz_id=str(quiz.quiz_id),
                    quiz_type=quiz.quiz_type,
                    quiz_title=quiz.title,
                    pass_score=quiz.pass_score,
                    score=score,
                    correct_count=correct,
                    total_count=quiz.question_count,
                    is_pass=score >= quiz.pass_score,
                    started_at=attempt.started_at,
                    submitted_at=attempt.submitted_at or attempt.started_at,
                )
            )
    return PaginationOut[QuizHistoryItemOut](
        items=items, total=total, page=page, size=size
    )


# ╔══════════════════════════════════════════════════════════════════╗
# ║ 4. 完成率报告 (UC-09)                                              ║
# ╚══════════════════════════════════════════════════════════════════╝


async def get_completion_report(quiz_id: int, *, current: UserSnapshot) -> QuizCompletionReportOut:
    """按院系统计完成率（管理员）。"""
    async with uow() as session:
        repo = QuizRepository(session)
        quiz = await repo.get_quiz(quiz_id)
        if quiz is None or quiz.quiz_type != QuizType.ASSIGNED:
            raise QuizNotFound()
        await _ensure_admin_can_access_quiz(session, current=current, quiz=quiz)

        # 目标学生集合 (user_id → dept_id)
        targets = await _resolve_quiz_recipients(
            session, target_scope=quiz.target_scope or {}
        )
        if await _is_dept_reviewer(session, current=current):
            targets = [u for u in targets if u.department_id == current.department_id]

        user_to_dept: dict[int, int] = {u.user_id: u.department_id for u in targets}
        dept_ids: set[int] = set(user_to_dept.values())

        # 已提交答题 (user_id → score)
        submitted = await repo.list_submitted_attempts_of_quiz_for_students(
            quiz_id=quiz_id, student_ids=list(user_to_dept.keys())
        )
        scored: dict[int, int] = {a.student_id: int(a.score or 0) for a in submitted}

        # 院系名字典
        dept_rows = (
            await session.execute(
                select(Department).where(Department.dept_id.in_(dept_ids or {0}))
            )
        ).scalars().all()
        dept_name_by_id = {d.dept_id: d.dept_name for d in dept_rows}

        # 按院系聚合
        by_dept: dict[int, dict[str, Any]] = {}
        for uid, did in user_to_dept.items():
            d = by_dept.setdefault(
                did,
                {"total": 0, "submitted": 0, "pass": 0, "score_sum": 0},
            )
            d["total"] += 1
            if uid in scored:
                d["submitted"] += 1
                d["score_sum"] += scored[uid]
                if scored[uid] >= quiz.pass_score:
                    d["pass"] += 1

        items = [
            DepartmentCompletionItem(
                dept_id=did,
                dept_name=dept_name_by_id.get(did, f"dept-{did}"),
                total_targets=v["total"],
                submitted_count=v["submitted"],
                completion_rate=_safe_div(v["submitted"], v["total"]),
                pass_count=v["pass"],
                pass_rate=_safe_div(v["pass"], v["submitted"]),
                avg_score=_safe_div(v["score_sum"], v["submitted"]),
            )
            for did, v in sorted(by_dept.items())
        ]

        total_targets = sum(v["total"] for v in by_dept.values())
        submitted_count = sum(v["submitted"] for v in by_dept.values())
        pass_count = sum(v["pass"] for v in by_dept.values())
        score_sum = sum(v["score_sum"] for v in by_dept.values())
        return QuizCompletionReportOut(
            quiz_id=str(quiz_id),
            title=quiz.title,
            status=quiz.status,
            deadline_at=quiz.deadline_at,
            total_targets=total_targets,
            submitted_count=submitted_count,
            completion_rate=_safe_div(submitted_count, total_targets),
            pass_rate=_safe_div(pass_count, submitted_count),
            avg_score=_safe_div(score_sum, submitted_count),
            by_department=items,
        )


async def export_completion_report_xlsx(
    quiz_id: int, *, current: UserSnapshot
) -> tuple[bytes, str]:
    """导出完成率报告为 XLSX（``bytes``, ``filename``）。"""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    report = await get_completion_report(quiz_id, current=current)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title="完成率报告")
    else:
        ws.title = "完成率报告"

    bold = Font(bold=True)
    # 总览
    ws.append(["测验 ID", report.quiz_id])
    ws.append(["标题", report.title])
    ws.append(["状态", report.status.value if hasattr(report.status, "value") else str(report.status)])
    ws.append(["截止时间", report.deadline_at.isoformat() if report.deadline_at else ""])
    ws.append(["目标人数", report.total_targets])
    ws.append(["已提交", report.submitted_count])
    ws.append(["完成率", f"{report.completion_rate * 100:.2f}%"])
    ws.append(["及格率", f"{report.pass_rate * 100:.2f}%"])
    ws.append(["平均分", f"{report.avg_score:.2f}"])
    ws.append([])

    headers = ["院系 ID", "院系", "目标人数", "已提交", "完成率", "及格人数", "及格率", "平均分"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold
    for d in report.by_department:
        ws.append(
            [
                d.dept_id,
                d.dept_name,
                d.total_targets,
                d.submitted_count,
                f"{d.completion_rate * 100:.2f}%",
                d.pass_count,
                f"{d.pass_rate * 100:.2f}%",
                f"{d.avg_score:.2f}",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_title = report.title.replace("/", "_").replace("\\", "_")[:40]
    filename = f"quiz_report_{quiz_id}_{safe_title}.xlsx"
    return buf.getvalue(), filename


# ╔══════════════════════════════════════════════════════════════════╗
# ║ 5. 内部工具                                                        ║
# ╚══════════════════════════════════════════════════════════════════╝


def _question_admin_out(q: QuestionBank) -> QuestionAdminOut:
    return QuestionAdminOut(
        question_id=str(q.question_id),
        content=q.content,
        option_a=q.option_a,
        option_b=q.option_b,
        option_c=q.option_c,
        option_d=q.option_d,
        correct_answer=q.correct_answer,
        explanation=q.explanation,
        fraud_type_id=q.fraud_type_id,
        knowledge_entry_id=str(q.knowledge_entry_id) if q.knowledge_entry_id else None,
        difficulty=q.difficulty,
        is_active=bool(q.is_active),
        created_by=str(q.created_by),
        created_at=q.created_at,
        updated_at=q.updated_at,
    )


def _quiz_detail_out(q: Quiz) -> QuizDetailOut:
    return QuizDetailOut(
        quiz_id=str(q.quiz_id),
        quiz_type=q.quiz_type,
        title=q.title,
        question_count=q.question_count,
        pass_score=q.pass_score,
        status=q.status,
        created_by=str(q.created_by),
        publish_level=int(getattr(q, "publish_level", 2) or 2),
        deadline_at=q.deadline_at,
        target_scope=q.target_scope,
        created_at=q.created_at,
    )


def _quiz_list_item_out(q: Quiz) -> QuizListItemOut:
    return QuizListItemOut(
        quiz_id=str(q.quiz_id),
        quiz_type=q.quiz_type,
        title=q.title,
        question_count=q.question_count,
        pass_score=q.pass_score,
        status=q.status,
        deadline_at=q.deadline_at,
        created_at=q.created_at,
        publish_level=int(getattr(q, "publish_level", 2) or 2),
    )


async def _is_dept_reviewer(session: AsyncSession, *, current: UserSnapshot) -> bool:
    role = await session.get(Role, current.role_id)
    return bool(role and role.role_code == "REVIEWER" and role.role_level == 1)


async def _is_school_reviewer(session: AsyncSession, *, current: UserSnapshot) -> bool:
    role = await session.get(Role, current.role_id)
    return bool(role and role.role_code == "REVIEWER" and role.role_level == 2)


# NOTE: legacy helper kept for now; access control uses quizzes.publish_level.
async def _is_dept_reviewer_user(session: AsyncSession, *, user_id: str) -> bool:
    role_level = (
        await session.execute(
            select(Role.role_level)
            .select_from(User)
            .join(Role, Role.role_id == User.role_id)
            .where(User.user_id == user_id)
        )
    ).scalar_one_or_none()
    return bool(role_level == 1)


async def _ensure_admin_can_access_quiz(
    session: AsyncSession, *, current: UserSnapshot, quiz: Quiz
) -> None:
    # 学校级审核员：看不到学院级（publish_level=1）发起的测验
    if await _is_school_reviewer(session, current=current):
        if int(getattr(quiz, "publish_level", 2) or 2) == 1:
            raise PermissionDenied("学校级审核员无权查看学院发布的测验")
        return

    # 院系级审核员：按范围可见（但撤销需要更严格的规则，见 _ensure_admin_can_cancel_quiz）
    if not await _is_dept_reviewer(session, current=current):
        return

    scope = quiz.target_scope or {}
    stype = scope.get("type")

    if quiz.created_by == current.user_id:
        return

    if stype == "ALL":
        return

    if stype == "DEPT":
        if current.department_id in set(scope.get("dept_ids") or []):
            return
        raise PermissionDenied("院系级审核员只能查看本院系测验")

    if stype == "USERS":
        user_ids = list(dict.fromkeys(scope.get("user_ids") or []))
        if not user_ids:
            raise PermissionDenied("测验范围非法")
        rows = (
            await session.execute(
                select(User.department_id)
                .join(Role, Role.role_id == User.role_id)
                .where(
                    User.user_id.in_(user_ids),
                    Role.role_code == "STUDENT",
                    User.status == UserStatus.ACTIVE.value,
                )
            )
        ).scalars().all()
        if rows and all(did == current.department_id for did in rows):
            return
        raise PermissionDenied("院系级审核员只能查看本院系测验")

    raise PermissionDenied("院系级审核员无权访问该测验")


async def _ensure_admin_can_cancel_quiz(
    session: AsyncSession, *, current: UserSnapshot, quiz: Quiz
) -> None:
    # 学校级审核员：仍然复用可见性规则
    if await _is_school_reviewer(session, current=current):
        await _ensure_admin_can_access_quiz(session, current=current, quiz=quiz)
        return

    # 院系级审核员：只能撤销自己发起的测验
    if await _is_dept_reviewer(session, current=current):
        if quiz.created_by != current.user_id:
            raise PermissionDenied("院系级审核员不能撤销学校发布的测验")
        return

    # 其它角色：权限码已在 API 层校验
    return


async def _build_target_scope(
    body: AssignedQuizCreateIn, *, session: AsyncSession
) -> dict[str, Any]:
    """把入参翻译为 ``target_scope`` JSON，并对 ID 做存在性校验。"""
    if body.scope_type == QuizScopeType.ALL:
        return {"type": "ALL"}
    if body.scope_type == QuizScopeType.DEPT:
        dept_ids = list(dict.fromkeys(body.dept_ids or []))
        if not dept_ids:
            raise QuizInvalidParam("按院系发起时必须提供 dept_ids")
        found = (
            (
                await session.execute(
                    select(Department.dept_id).where(Department.dept_id.in_(dept_ids))
                )
            )
            .scalars()
            .all()
        )
        if len(found) != len(dept_ids):
            raise QuizInvalidParam("存在未找到的院系 ID")
        return {"type": "DEPT", "dept_ids": dept_ids}
    # USERS
    user_ids = list(dict.fromkeys(body.user_ids or []))
    if not user_ids:
        raise QuizInvalidParam("按学生发起时必须提供 user_ids")
    found_users = (
        await session.execute(
            select(User.user_id)
            .join(Role, Role.role_id == User.role_id)
            .where(
                User.user_id.in_(user_ids),
                Role.role_code == "STUDENT",
                User.status == UserStatus.ACTIVE.value,
            )
        )
    ).scalars().all()
    if set(found_users) != set(user_ids):
        raise QuizInvalidParam("user_ids 包含非学生或非启用账户")
    return {"type": "USERS", "user_ids": user_ids}


async def _resolve_quiz_recipients(
    session: AsyncSession, *, target_scope: dict[str, Any]
) -> list[User]:
    """解析参与学生（ACTIVE + STUDENT 角色）。"""
    stype = target_scope.get("type") if target_scope else None
    stmt = (
        select(User)
        .join(Role, Role.role_id == User.role_id)
        .where(
            Role.role_code == "STUDENT",
            User.status == UserStatus.ACTIVE.value,
        )
    )
    if stype == "ALL" or stype is None:
        pass
    elif stype == "DEPT":
        dept_ids = target_scope.get("dept_ids") or []
        if not dept_ids:
            return []
        stmt = stmt.where(User.department_id.in_(dept_ids))
    elif stype == "USERS":
        user_ids = target_scope.get("user_ids") or []
        if not user_ids:
            return []
        stmt = stmt.where(User.user_id.in_(user_ids))
    else:
        return []
    return list((await session.execute(stmt)).scalars().all())


async def _ensure_quiz_open(quiz: Quiz) -> None:
    """校验 ASSIGNED 测验仍可作答（ACTIVE + 未过截止）。"""
    if quiz.status != QuizStatus.ACTIVE:
        raise QuizClosed()
    if quiz.deadline_at is not None:
        # quiz.deadline_at 是 naive UTC（DB 默认），统一按 UTC 比较
        deadline = quiz.deadline_at
        if deadline.tzinfo is None:
            deadline = deadline.replace(tzinfo=UTC)
        if deadline <= datetime.now(tz=UTC):
            raise QuizClosed("测验已超过截止时间")


async def _ensure_student_in_scope(
    session: AsyncSession, *, quiz: Quiz, user_id: int
) -> None:
    """校验当前学生在 quiz.target_scope 内。"""
    scope = quiz.target_scope or {}
    stype = scope.get("type")
    if stype == "ALL":
        return
    if stype == "USERS":
        if user_id in set(scope.get("user_ids") or []):
            return
        raise QuizNotEligible()
    if stype == "DEPT":
        user = await session.get(User, user_id)
        if user is None:
            raise QuizNotEligible()
        if user.department_id in set(scope.get("dept_ids") or []):
            return
        raise QuizNotEligible()
    raise QuizNotEligible()


def _safe_div(a: int | float, b: int | float) -> float:
    if not b:
        return 0.0
    return float(a) / float(b)


# ╔══════════════════════════════════════════════════════════════════╗
# ║ 6. 定时任务：截止前 24h 提醒                                       ║
# ╚══════════════════════════════════════════════════════════════════╝


async def send_quiz_deadline_reminders() -> int:
    """扫描 24h 内到期的 ACTIVE 测验，向未提交的学生发提醒。

    返回累计发出的通知条数。每个测验只发一次（``reminder_sent=1``）。
    """
    now = datetime.now(tz=UTC)
    window_end = now + timedelta(hours=24)
    total_sent = 0
    async with uow() as session:
        repo = QuizRepository(session)
        quizzes = await repo.list_quizzes_near_deadline(
            window_start=now, window_end=window_end
        )
        for quiz in quizzes:
            recipients = await _resolve_quiz_recipients(
                session, target_scope=quiz.target_scope or {}
            )
            submitted = await repo.list_submitted_attempts_of_quiz(quiz.quiz_id)
            submitted_ids = {a.student_id for a in submitted}
            for stu in recipients:
                if stu.user_id in submitted_ids:
                    continue
                await send_notification(
                    recipient_id=stu.user_id,
                    type="QUIZ_DEADLINE_REMINDER",
                    title=f"[测验提醒] {quiz.title}",
                    content=(
                        f"该测验将在 {quiz.deadline_at.isoformat() if quiz.deadline_at else ''}"
                        f" 截止，请尽快完成"
                    ),
                    related_object_type="quiz",
                    related_object_id=quiz.quiz_id,
                    db_session=session,
                )
                total_sent += 1
            quiz.reminder_sent = 1
            await session.flush()
    if total_sent:
        logger.info("quiz_deadline_reminders_sent", count=total_sent)
    return total_sent